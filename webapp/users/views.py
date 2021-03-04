from django.shortcuts import render, redirect, HttpResponseRedirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.models import User
from django.contrib import messages
from .forms import NewUserForm, OpportunityBuyingForm, OpportunitySellingForm, AgentProposalForm, FeaturesForm, ProposalForm, AgentForm, ImageForm, UserEdit, PostingImageBuyingForm, PostingImageSellingForm, Promocode
from .models import House, OpportunitieBuying, OpportunitieSelling, Proposals ,LocationBuying, LocationSelling, Features, Condition, RealtorQualities, Realtor, RealtorImages, PostingImageBuying, PostingImageSelling, BlogPost, GeoLocation
from django.db.models import Max
import random
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from .notifications import my_handler
from notifications.signals import notify
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
import stripe
from templated_email import send_templated_mail
from django.contrib.auth.decorators import user_passes_test
from mailchimp3 import MailChimp
from .utils import SendSubscribeMailAgent, SendSubscribeMailUser

from django.db import connection

#USERS

stripe.api_key = settings.STRIPE_SECRET_KEY

def user_check(user):
    return user.groups.filter(name='buyer').exists()

def agent_check(user):
    return user.groups.filter(name='agent').exists()


# TODO: Limit access based on subscription status
def agent_sub_check(user):
    if user.is_active:
        
        try:
            go = Realtor.objects.get(user=user)
            sub = go.subscription_id
        except Realtor.DoesNotExist:
            sub = ''
    print(sub)
    return sub != ''


def legal(request):
    return render(request = request, 
                  template_name='users/legal.html', 
                  
                  )


def new_opp_pg1(request):
    current_user = request.user
    return render(request = request, 
                  template_name='form/new_opp_pg1.html', 
                  
                  )
#@user_passes_test(agent_sub_check, login_url='/profile_edit')
@login_required
def dashboard(request):
    user = request.user
    blog = BlogPost.objects.all().order_by('created_date').reverse()
    #------------------------------------------------------------------------
    # update users_heading table. The values in kwargs need to be soft-coded based on the current user info
    # example below. uncomment it and edit kwargs and see effect
#    kwargs = {  'first_name' : 'Jerry',
#                'last_name' : 'Wang',
#                'email' : 'jerrywang@gmail.com',
#                'role' : 'Seller',
#                'company_name' : 'Jerry\'\'s',
#                }
#    tbl = 'users_heading'
#    column_for_count = 'user_name'
#    update_table(request, tbl, column_for_count, **kwargs)
    #------------------------------------------------------------------------
    return render(request = request, 
                  template_name='users/user/dashboard.html',
                  context= {"blog":blog})

#=============================================================================
# pass in the fields that needs to be updated in the format of a dict.
# ex: d = {'username':'Tom', 'email':'hello@gmail.com'}
# Please see def dashboard(request) for example
UPDATE_TBL_SQL = """UPDATE {tbl}
                    SET {fltr}
                    WHERE user_name='{user_name}'
                    """
INSERT_SQL = """INSERT INTO {tbl} ({fltr_key},user_name)
                VALUES ({fltr_value},'{user_name}')
                """
# for sqlite3
FIELDS_SQL = """    PRAGMA table_info('{tbl}')
                """
# for postgresql
#FIELDS_SQL = """SELECT column_name
#                FROM information_schema.columns
#                WHERE table_name='{tbl}'
#                """
def update_table(request, tbl, column_for_count, **kwargs):
    # parse request
    user = request.user
    print (user)
    # parse data
#    row, created = User.objects.get_or_create(id=id)
    cursor = connection.cursor()
    fields_list = []
    # for sqlite3
    [fields_list.append(r[1]) for r in cursor.execute(FIELDS_SQL.format(tbl=tbl)).fetchall()]
    # for postgresql
#    [fields_list.append(r[0]) for r in cursor.execute(FIELDS_SQL.format(tbl=tbl)).fetchall()]
    if int(cursor.execute("""   SELECT COUNT({column_for_count}) AS 'cnt'
                                FROM {tbl}
                                WHERE user_name='{user_name}'
                            """.format(column_for_count=column_for_count,
                                       user_name=user,
                                       tbl=tbl)).fetchone()[0]) != 0:
#        id = int(list(request.session.values())[0])
        cursor.execute(UPDATE_TBL_SQL.format(  tbl=tbl,
                                               user_name=user,
                                               fltr= ','.join(["""%s='%s'""" % (k,v) for k, v in kwargs.items() if k in fields_list])
                                               )
                       )
    else:
        cursor.execute(INSERT_SQL.format(tbl=tbl,
                                        fltr_key=','.join(["""%s""" % (k) for k, v in kwargs.items() if k in fields_list]),
                                        fltr_value=','.join(["""'%s'""" % (v) for k, v in kwargs.items() if k in fields_list]),
                                        column_for_count=column_for_count,
                                        user_name=user))
#=============================================================================
                  
def start(request):
    return render(request = request, 
                  template_name='users/start.html')                  

@login_required
def my_opportunities(request):
    current_user = request.user
    buying = OpportunitieBuying.objects.filter(user_id=current_user.id)
    selling = OpportunitieSelling.objects.filter(user_id=current_user.id)
    return render(request=request,
                  template_name='users/user/my_opportunities.html',
                  context= {"buying":buying,
                            "selling":selling,
                            })


@login_required
def new_buying(request):
    if request.method == 'POST':
        form = OpportunityBuyingForm(request.POST)
        if form.is_valid():              
            current_user = request.user
            
            posting_id = int(request.session["current_posting"])
            print(posting_id)
            posting = OpportunitieBuying.objects.get(pk=posting_id) 
            posting.bedrooms = form.cleaned_data['bedrooms']
            posting.bathrooms = form.cleaned_data['bathrooms']
            posting.budget = form.cleaned_data['budget']
            posting.sq_ft = form.cleaned_data['sq_ft']
            posting.looking_for = form.cleaned_data['looking_for']
            posting.notes_to_agent = form.cleaned_data['notes_to_agent']
            posting.notes_about = form.cleaned_data['notes_about']
            posting.insite1 = form.cleaned_data['insite1']
            posting.insite2 = form.cleaned_data['insite2']
            posting.insite3 = form.cleaned_data['insite3']
            posting.insite4 = form.cleaned_data['insite4']
            posting.urgency = form.cleaned_data['urgency']

            features = form.cleaned_data.get("features") 
            condition = form.cleaned_data.get("condition") 
            realtorqualities = form.cleaned_data.get("realtorqualities") 

            posting.features.set(features)
            posting.condition.set(condition)
            posting.realtorqualities.set(realtorqualities)
            posting.save()
            #need to add city filter
            notify.send(
                        current_user, 
                        recipient=User.objects.filter(groups__name='agent'), 
                        verb='is looking to buy a house', 
                        description="buying",action_object=OpportunitieBuying.objects.get(pk=posting.pk))

            return redirect('users:buy_posting_view', id=posting.pk)
          
        else:
            print("nope")
            return render(request = request,
                          template_name = "users/user/new_buying_opp.html",
                          context={"form":form})
    form = OpportunityBuyingForm
    
    formFeatures = FeaturesForm
    return render(request = request,
                  template_name = "users/user/new_buying_opp.html",
                  context={"form":form,
                            "features":formFeatures})
@login_required
def add_images_buy(request, id):
    if request.method == 'POST':
        
        form = PostingImageBuyingForm(request.POST, request.FILES)
        if form.is_valid():  
            posting = OpportunitieBuying.objects.get(pk=id) 
            image_save = form.save(commit = False)
            image_save.posting = posting
            image_save.save()
            return redirect('users:buy_posting_view', id=posting.pk)
    form = PostingImageBuyingForm
    return render(request = request,
                  template_name = "users/user/new_buying_image.html",
                  context={"form":form})


@login_required
def new_buying_location(request):
    if request.method == 'POST':
        new_posting = OpportunitieBuying()
        
        new_posting_location = LocationBuying()
        new_posting_location.city = request.POST.get('city')
        new_posting_location.neighbourhood = request.POST.get('neighbourhood')
        new_posting_location.province = request.POST.get('province')
        new_posting_location.country = request.POST.get('country')
        new_posting_location.save()
        location = LocationBuying.objects.get(pk=new_posting_location.pk)

        current_user = request.user
        new_posting.user_id = current_user
        new_posting.location = location
        new_posting.save()
        
        request.session['current_posting'] = str(new_posting.pk)
        return redirect("users:buying")


    return render(request = request,
                  template_name = "users/user/new_buying_opp_location.html")

@login_required
def new_selling(request):
    if request.method == 'POST':
        form = OpportunitySellingForm(request.POST)
        if form.is_valid():
            current_user = request.user
            posting_id = int(request.session["current_posting"])
            posting = OpportunitieSelling.objects.get(pk=posting_id)
            posting.bedrooms = form.cleaned_data['bedrooms']
            posting.bathrooms = form.cleaned_data['bathrooms']
            posting.sq_ft = form.cleaned_data['sq_ft']
            posting.year_built = form.cleaned_data['year_built']
            posting.asking_price = form.cleaned_data['asking_price']
            posting.unique_value = form.cleaned_data['unique_value']
            posting.details = form.cleaned_data['details']
            posting.love = form.cleaned_data['love']
            posting.notes_to_agent = form.cleaned_data['notes_to_agent']
            posting.phone = form.cleaned_data['phone']
            posting.urgency = form.cleaned_data['urgency']

            features = form.cleaned_data.get("features") 
            realtorqualities = form.cleaned_data.get("realtorqualities") 
            building_types = form.cleaned_data.get("building_types") 
            posting.features.set(features)
            posting.realtorqualities.set(realtorqualities)
            posting.building_types.set(building_types)
            posting.save()
            notify.send(
                        current_user, 
                        recipient=User.objects.filter(groups__name='agent'), 
                        verb='is looking to sell a house', 
                        description="selling",action_object=OpportunitieSelling.objects.get(pk=posting.pk))

            return redirect('users:sell_posting_view', id=posting.pk)
        else:
            return render(request = request,
                  template_name = "users/user/new_selling_opp.html",
                  context={"form":form})
        
    form = OpportunitySellingForm
    return render(request = request,
                  template_name = "users/user/new_selling_opp.html",
                  context={"form":form})

@login_required
def add_images_sell(request, id):
    if request.method == 'POST':
        form = PostingImageSellingForm(request.POST, request.FILES)
        if form.is_valid():  
            posting = OpportunitieSelling.objects.get(pk=id) 
            image_save = form.save(commit = False)
            image_save.posting = posting
            image_save.save()
            return redirect('users:buy_posting_view', id=posting.pk)
    form = PostingImageSellingForm
    return render(request = request,
                  template_name = "users/user/new_selling_image.html",
                  context={"form":form})

@login_required
def new_selling_location(request):
    if request.method == 'POST':
        new_posting = OpportunitieSelling()
        
        new_posting_location = LocationSelling()
        new_posting_location.street_number = request.POST.get('street_number')
        new_posting_location.route = request.POST.get('route')
        new_posting_location.city = request.POST.get('city')
        new_posting_location.neighbourhood = request.POST.get('neighbourhood')
        new_posting_location.province = request.POST.get('province')
        new_posting_location.country = request.POST.get('country')
        new_posting_location.save()
        location = LocationSelling.objects.get(pk=new_posting_location.pk)

        current_user = request.user
        new_posting.user_id = current_user
        new_posting.location = location
        new_posting.save()
        
        request.session['current_posting'] = str(new_posting.pk)
        #post_save.connect(my_handler, sender=current_user)
        
        return redirect("users:selling")


    return render(request = request,
                  template_name = "users/user/new_selling_opp_location.html")

@login_required
def user_proposals(request):
    current_user = request.user
    proposals_buy = Proposals.objects.filter(buying__user_id=current_user)
    proposals_sell = Proposals.objects.filter(selling__user_id=current_user)

    return render(request = request, 
                  template_name='users/user/proposals.html',
                  context={"buy":proposals_buy, "sell":proposals_sell})   

@login_required
def buy_posting_view(request, id):
    posting = OpportunitieBuying.objects.get(pk=id)
    try:
        image = PostingImageBuying.objects.filter(posting__pk=id)
    except ObjectDoesNotExist:
        image = None
    return render(request = request, 
                  template_name='users/user/buying_posting_single.html',
                  context={"posting":posting,"image":image}
                  )

@login_required
def post_edit_buy(request, id):
    posting = get_object_or_404(OpportunitieBuying, pk=id)
    if request.method == "POST":
        form = OpportunityBuyingForm(request.POST, instance=posting)
        if form.is_valid():
            posting = form.save(commit=False)
            posting.save()
            return redirect('users:buy_posting_view', id=posting.pk)
    else:
        form = OpportunityBuyingForm(instance=posting)
    return render(request = request, 
                  template_name='users/user/buying_posting_edit.html',
                  context={'form':form, 'posting':posting}
                  )

@login_required
def sell_posting_view(request, id):
    posting = OpportunitieSelling.objects.get(pk=id)
    try:
        image = PostingImageSelling.objects.filter(posting__pk=id)
    except ObjectDoesNotExist:
        image = None
    return render(request = request, 
                  template_name='users/user/selling_posting_single.html',
                  context={"posting":posting,"image":image}
                  )

@login_required
def post_edit_sell(request, id):
    posting = get_object_or_404(OpportunitieSelling, pk=id)
    if request.method == "POST":
        form = OpportunitySellingForm(request.POST, instance=posting)
        if form.is_valid():
            posting = form.save(commit=False)
            posting.save()
            return redirect('users:sell_posting_view', id=posting.pk)
    else:
        form = OpportunitySellingForm(instance=posting)
    return render(request = request, 
                  template_name='users/user/selling_posting_edit.html',
                  context={'form':form}
                  )

@login_required
def single_proposal(request, id):
    prop = Proposals.objects.get(pk=id)
    agent = Realtor.objects.get(user__pk=prop.agent.pk)
    img = get_object_or_404(RealtorImages, user=prop.agent.pk)
    print(agent.work_phone)
    return render(request = request, 
                  template_name='users/user/proposal_single.html',
                  context={"proposal":prop, "agent":agent, "img":img}
                  )

@login_required
def accept_proposal(request, id):
    proposal = Proposals.objects.get(pk=id)
    proposal.accepted_user = True
    proposal.save()
    current_user = request.user
    notify.send(
                current_user, 
                recipient=User.objects.get(pk=proposal.agent.pk), 
                verb='accepted your proposal', 
                description="proposal",
                action_object=proposal)
    send_templated_mail(
                            template_name='step2_proposal',
                            from_email='info@listingllama.com',
                            recipient_list=[request.user.email],
                            context={
                                'username':request.user.username,
                                'full_name':request.user.get_full_name(),
                                'signup_date':request.user.date_joined
                            },)
    return redirect('users:single_proposal', id=proposal.pk)

@login_required
def accept_proposal_agent(request, id):
    proposal = Proposals.objects.get(pk=id)
    proposal.accepted_agent = True
    proposal.save()
    if proposal.buying:
        user_id = proposal.buying.user_id.pk
    else:
        user_id = proposal.selling.user_id.pk
    current_user = request.user
    notify.send(
                current_user, 
                recipient=User.objects.get(pk=user_id), 
                verb='would like to work with you!', 
                description="proposal",
                action_object=proposal)
    send_templated_mail(
                            template_name='step2_proposal',
                            from_email='info@listingllama.com',
                            recipient_list=[request.user.email],
                            context={
                                'username':request.user.username,
                                'full_name':request.user.get_full_name(),
                                'signup_date':request.user.date_joined
                            },)            
    return redirect('users:single_proposal', id=proposal.pk)
    
@login_required
def realtor_find(request):
    max_id = User.objects.all().aggregate(max_id=Max("id"))['max_id']
    while True:
        pk = random.randint(1, max_id)
        realtor = User.objects.filter(groups__name='agent',pk=pk).first()
        
        if realtor:
            profile = Realtor.objects.get(user=realtor)
            if profile.company != '' or None and profile.about_me != '' or None:
                img = get_object_or_404(RealtorImages, user=realtor)
                return render(request = request, 
                    template_name='users/user/realtor_find.html',
                    context={"realtor":realtor, "profile":profile, "img":img}
                    )

@login_required
def realtor_find_next(request, id):
    request.session["last"] = id
    qty = User.objects.last()
    print(qty.pk)
    pk = id
    while True:
        pk += 1
        if pk > qty.pk:
            pk = 1
        print(pk)
        realtor = User.objects.filter(groups__name='agent',pk=pk).first()
        if realtor:
            profile = Realtor.objects.get(user=realtor)
            if profile.company != None and profile.about_me != None:
                img = get_object_or_404(RealtorImages, user=realtor)
                return render(request = request, 
                    template_name='users/user/realtor_find.html',
                    context={"realtor":realtor, "profile":profile, "img":img}
                    )

@login_required
def realtor_find_back(request):
    pk = request.session["last"]
    while True:
        realtor = User.objects.filter(groups__name='agent',pk=pk).first()
        if realtor:
            profile = Realtor.objects.get(user=realtor)
            img = get_object_or_404(RealtorImages, user=realtor)
            return render(request = request, 
                  template_name='users/user/realtor_find.html',
                  context={"realtor":realtor, "profile":profile, "img":img}
                  )



#AGENT

@login_required
def payment_page(request):
    if request.method == 'POST':
        form = Promocode(request.POST)
        if form.is_valid():
            request.session['promo'] = form.cleaned_data['promo']
            return redirect("users:charge")
    else:
        form = Promocode()
    return render(request = request, 
                  template_name='users/promo.html',
                  context={"form":form})


def charge(request): # new
    form = Promocode()
    if request.method == 'POST':
        promo = request.session.get('promo')

        customer = stripe.Customer.create(
                email=request.POST['stripeEmail'],
        )
        current_user = request.user
        realtor = get_object_or_404(Realtor, user=current_user)
        realtor.strip_id = customer.id
        
        charge = stripe.Subscription.create(
            customer=customer,
            trial_period_days= 14,   
            items = [{
                    'plan': settings.STRIPE_PLAN,#plan_FOAgwGMWaIzQEM
            }],
            source=request.POST['stripeToken'],
            coupon=promo,
        )
        print(charge.id)
        realtor.subscription_id = charge.id
        realtor.save()
        SendSubscribeMailAgent(request.user) # Send the Mail, Class available in utils.py
        return redirect("users:profile_edit")
    return render(request = request, 
                  template_name='users/payment.html',
                  context={"key":settings.STRIPE_PUBLISHABLE_KEY, "form":form})

def cancel_sub(request):
    current_user = request.user
    realtor = get_object_or_404(Realtor, user=current_user)
    cancel = stripe.Subscription.delete(realtor.subscription_id)
    realtor.subscription_id = ''
    realtor.save()
    return redirect("users:profile_edit")

@login_required
def profile_edit(request):
    current_user = request.user
    posting = get_object_or_404(Realtor, user=current_user)
    cur_user = get_object_or_404(User, pk=current_user.pk)
    profile_img = get_object_or_404(RealtorImages, user=current_user)
    if request.method == "POST":
        form = AgentForm(request.POST, instance=posting)
        form2 = UserEdit(request.POST)
        if form.is_valid() and form2.is_valid():
            cur_user.email = form2.cleaned_data['email']
            cur_user.first_name = form2.cleaned_data['first_name']
            cur_user.last_name = form2.cleaned_data['last_name']
            posting = form.save(commit=False)
            posting.save()
            form.save_m2m()
            cur_user.save()
      
            return redirect('users:profile_edit')
    else:
        form = AgentForm(instance=posting)
        form2 = NewUserForm(instance=cur_user)
    return render(request = request, 
                  template_name='users/agent/profile_edit.html',
                  context={'form':form, 'form2':form2, "profile_img":profile_img, 'STATIC_URL': settings.STATIC_URL, 'status':posting}
                  )

@login_required
def agent_profile(request):
    form = AgentForm
    if request.method == 'POST':
        form = AgentForm(request.POST)
        if form.is_valid():
            obj = Realtor.objects.get(user=request.user)
            obj.company = form.cleaned_data['company']
            obj.work_phone = form.cleaned_data['work_phone']
            obj.cell_phone = form.cleaned_data['cell_phone']
            obj.about_me = form.cleaned_data['about_me']
            obj.why_work = form.cleaned_data['why_work']
            services = form.cleaned_data.get("services")
            obj.services.set(services)
            cities = form.cleaned_data.get('cities')
            obj.cities.set(cities)
            special = form.cleaned_data.get('special')
            obj.special.set(special)
            languages = form.cleaned_data.get('languages')
            obj.languages.set(languages)
            building_types = form.cleaned_data.get('building_types')
            obj.building_types.set(building_types)
            obj.website = form.cleaned_data['website']
            obj.save()
            return redirect("users:dashboard")

    return render(request = request, 
                    template_name='users/agent/profile.html',
                    context= {"form":form})   


@login_required
def upload_picture(request):
    form = ImageForm()
    current_user = request.user
    if RealtorImages.objects.get(user=current_user):
        img = RealtorImages.objects.get(user=current_user)
    if request.method == 'POST':
        form = ImageForm(request.POST, request.FILES)
        if form.is_valid():
            
            if RealtorImages.objects.get(user=current_user):
                upload = RealtorImages.objects.get(user=current_user)
                if request.FILES.get('profile'):
                    upload.profile.delete()
                    upload.profile = request.FILES.get('profile')
                if request.FILES.get('cover'):
                    upload.cover.delete()
                    upload.cover = request.FILES.get('cover')
                upload.save()
            else:
                upload = RealtorImages()
                upload.user = current_user
                upload.profile = request.FILES.get('profile')
                upload.cover = request.FILES.get('cover')
                upload.save()
            
            return redirect("users:profile_edit")
    return render(request = request, 
                  template_name='users/agent/image.html',
                  context={"form":form, "img":img})

@login_required
def agent_dashboard(request):
    return render(request = request, 
                  template_name='users/agent/dashboard.html',
                  context= {"opprotunities":OpportunitieBuying.objects.all()})    
@login_required                  
def agent_opp(request):
    current_user = request.user
    posting = get_object_or_404(Realtor, user=current_user)
    if posting.subscription_id == '' or posting.subscription_id == None:
        return redirect("users:profile_edit")
    else:
        props_sell = Proposals.objects.values_list('selling', flat=True).filter(agent=current_user).filter(buying=None)
        props_buy = Proposals.objects.values_list('buying', flat=True).filter(agent=current_user).filter(selling=None)
        buying_opp = OpportunitieBuying.objects.exclude(pk__in=props_buy)
        selling_opp = OpportunitieSelling.objects.exclude(pk__in=props_sell)
        print(buying_opp)
        return render(request = request, 
                    template_name='users/agent/opportunities.html',
                    context= {"buying":buying_opp,
                                "selling":selling_opp})
@login_required
def agent_proposal_buying(request):
    current_user = str(request.user)
    request.session[current_user] = request.POST.get("opp_id")
    request.session['opp_type'] = "buying"
    return redirect("users:new_proposals")

@login_required
def agent_proposal_selling(request):
    current_user = str(request.user)
    request.session[current_user] = request.POST.get("opp_id")
    request.session['opp_type'] = "selling"
    return redirect("users:new_proposals")   

@login_required
def agent_proposals(request):
    current_user = request.user
    posting = get_object_or_404(Realtor, user=current_user)
    if posting.subscription_id == '' or posting.subscription_id == None:
        return redirect("users:profile_edit")
    else:

        return render(request = request, 
                    template_name='users/agent/proposals.html',
                    context= {"prop":Proposals.objects.filter(agent=current_user.id),
                                "buying":OpportunitieBuying.objects.all()})
@login_required
def agent_create_proposal(request):
    form = ProposalForm
    if request.method == 'POST':
        form = ProposalForm(request.POST)
        if form.is_valid():
            current_user = request.user
            opp_id = request.session[str(current_user)]
            obj = Proposals()
            obj.agent = current_user
            obj.save()
            obj.subject = form.cleaned_data['subject']
            obj.note = form.cleaned_data['note']

            services = form.cleaned_data.get("services")
            obj.services.set(services)

            obj.fee = form.cleaned_data['fee']
            if request.session['opp_type'] == "buying":
                obj.buying = OpportunitieBuying.objects.get(pk=opp_id) 
                user = OpportunitieBuying.objects.get(pk=opp_id) 
                user = user.user_id
                notify.send(
                            current_user, 
                            recipient=User.objects.get(pk=user.pk), 
                            verb='sent you a proposal!', 
                            description="proposal",
                            action_object=Proposals.objects.get(pk=obj.pk))
                send_templated_mail(
                            template_name='step1_proposal',
                            from_email='info@listingllama.com',
                            recipient_list=[request.user.email],
                            context={
                                'username':request.user.username,
                                'full_name':request.user.get_full_name(),
                                'signup_date':request.user.date_joined
                            },)
            else:
                obj.selling = OpportunitieSelling.objects.get(pk=opp_id) 
                user = OpportunitieBuying.objects.get(pk=opp_id) 
                user = user.user_id
                notify.send(
                            current_user, 
                            recipient=User.objects.get(pk=user.pk), 
                            verb='sent you a proposal!', 
                            description="proposal",
                            action_object=Proposals.objects.get(pk=obj.pk))

                send_templated_mail(
                            template_name='step1_proposal',
                            from_email='info@listingllama.com',
                            recipient_list=[request.user.email],
                            context={
                                'username':request.user.username,
                                'full_name':request.user.get_full_name(),
                                'signup_date':request.user.date_joined
                            },)
            obj.save()
            del request.session[str(current_user)]
            
            return redirect("users:agent_proposals")

    return render(request = request, 
                    template_name='users/agent/new_proposal.html',
                    context= {"form":form})   

def create_user(strategy, backend, details, user=None, *args, **kwargs):
    if strategy.session_get('user_type'):   
        group_name = strategy.session_get('user_type')
        group = Group.objects.get(name=group_name)
        user.groups.add(group)
        if group_name == 'agent':
            profile = RealtorImages()
            profile.user = User.objects.get(pk=user.pk)
            profile.profile = "images/Profile-Placeholder-Logo-DkGrey.png"
            profile.cover = "images/GreyBG.jpg"
            profile.save()
            obj = Realtor()
            obj.user = User.objects.get(pk=user.pk)
            obj.save()
            SendSubscribeMailAgent(request.user)
        else:
            SendSubscribeMailUser(request.user)
            print("done")




def register(request, backend='django.contrib.auth.backends.ModelBackend'):
    if request.method == "POST":
        form = NewUserForm(request.POST)
        if form.is_valid():
            
            user = form.save()
            user.backend = 'django.contrib.auth.backends.ModelBackend'
            username = form.cleaned_data.get('username')
            group_name = form.cleaned_data.get('user_type')
            group = Group.objects.get(name=group_name)
            user.groups.add(group)
            messages.success(request, f"New account created: {username}")
            login(request, user)
            if group_name == 'agent':
                profile = RealtorImages()
                profile.user = User.objects.get(pk=request.user.pk)
                profile.profile = "images/Profile-Placeholder-Logo-DkGrey.png"
                profile.cover = "images/GreyBG.jpg"
                profile.save()
                obj = Realtor()
                obj.user = User.objects.get(pk=request.user.pk)
                obj.save()
                SendSubscribeMailAgent(request.user)
                return redirect("users:promo")
            else:
                SendSubscribeMailUser(request.user) # Send the Mail, Class available in utils.py
            return redirect("users:dashboard")

        else:
            for msg in form.error_messages:
                messages.error(request, f"{msg}: {form.error_messages[msg]}")

            return render(request = request,
                          template_name = "users/register.html",
                          context={"form":form})

    form = NewUserForm
    return render(request = request,
                  template_name = "users/register.html",
                  context={"form":form})

def login_request(request):
    if request.method == 'POST':
        form = AuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.info(request, f"You are now logged in as {username}")
                return redirect('/')
            else:
                messages.error(request, "Invalid username or password.")
        else:
            messages.error(request, "Invalid username or password.")
    form = AuthenticationForm()
    return render(request = request,
                    template_name = "users/login.html",
                    context={"form":form})

def logout_request(request):
    logout(request)
    messages.info(request, "Logged out successfully!")
    return redirect("users:homepage")

#=====================================================================================
#-------------------------------------------------------------------------------------
# Google map API does not show all cities in radius, the problem varies.
# Calculation is handled by google api.
#-------------------------------------------------------------------------------------
import urllib
from urllib.request import urlopen
import json
GOOGLE_MAP_API_KEY = 'AIzaSyDQIGA7Ic7jyuWpzXPwKkDZY7ZYMYAvizk'
SEARCH_TYPES = ('locality',
#                'political',
#                'archipelago',
#                'colloquial_area',
                'administrative_area_level_1',
                'administrative_area_level_2',
                'administrative_area_level_3',
                'administrative_area_level_4',
                'administrative_area_level_5',
#                'sublocality',
#                'sublocality_level_1',
#                'sublocality_level_2',
#                'sublocality_level_3',
#                'sublocality_level_4',
#                'sublocality_level_5',
#                'cities',   #this returns hotels...weird
                )
# For example:
# city = 'Kelowna' or 'red deer' (any space is ok)
# province = 'AB'
# country = 'CA'
# radius = 50000 or '50000' (whatever str or int or float)
def cities_in_radius(city, province, country, radius):
    address = ''.join(replace(city).split(' ')) + ',' + province + ',' + country
    url = 'https://maps.googleapis.com/maps/api/geocode/json?address={address}&key={key}'.format(address=address,
                                                                                                 key=GOOGLE_MAP_API_KEY)
    location = ','.join([str(r) for r in list(request_response(url)['results'][0]['geometry']['location'].values())])
    print (location)
    cities = []
    for type in SEARCH_TYPES:
        url = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={location}&radius={radius}&type={type}&key={key}'.format(location=location,
                                                                                                                                              radius=str(radius),
                                                                                                                                              type=type,
                                                                                                                                              key=GOOGLE_MAP_API_KEY)
        [cities.append(r['name']) for r in request_response(url)['results']]
    return list(set(cities))

def request_response(url):
    try:
        response = urlopen(url).read().decode('utf-8')
        responseJson = json.loads(response)
        return responseJson
    except Exception as ex:
        print (ex)

def replace(s):
    return s.replace("'", "\'").replace('"', '\"')

def _flt(v, dflt=None):
    try:
        return float(v)
    except:
        return dflt

#=====================================================================================
#-------------------------------------------------------------------------------------
# geonames provides all cities info for a country. Do calculation by myself.
#-------------------------------------------------------------------------------------
import sqlite3
import os
import math
# timezone must be one of the following:
#'Atikokan', 'Blanc-Sablon', 'Cambridge_Bay', 'Chicago', 'Creston', 'Dawson', 'Dawson_Creek', 'Denver',
#'Detroit', 'Easter', 'Edmonton', 'Fort_Nelson', 'Glace_Bay', 'Godthab', 'Goose_Bay', 'Halifax', 'Inuvik',
#'Iqaluit', 'Juneau', 'London', 'Los_Angeles', 'Menominee', 'Miquelon', 'Moncton', 'New_York', 'Nipigon',
#'Pangnirtung', 'Rainy_River', 'Rankin_Inlet', 'Regina', 'Resolute', 'Sitka', 'St_Johns', 'Swift_Current',
#'Thule', 'Thunder_Bay', 'Toronto', 'Vancouver', 'Whitehorse', 'Winnipeg', 'Yellowknife'
CITIES_IN_DISTANCE_SQL = """SELECT name, distance
                            FROM (SELECT *, ROUND( SQRT( POWER( ( (69.1/1.61) * ('{lat}' - latitude)), 2) + POWER(( (53/1.61) * ('{lng}' - longitude)), 2)), 1) AS distance
                                    FROM {tbl}
                                    ) d
                            WHERE d.distance<{radius}
                                AND d.feature_code='PPL'
                            ORDER BY d.distance DESC
                            """
FETCH_LAT_LNG_SQL = """ SELECT name, latitude, longitude
                        FROM {tbl}
                        WHERE name='{city}'
                            AND feature_code='PPL'
                            AND timezone LIKE '%{timezone}%'
                        """
def cities_in_radius_2(city, timezone, radius):
    tbl='users_geolocation'
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # django built-in connection
#    cursor = connection.cursor()
    # sqlite3
    con = sqlite3.connect(BASE_DIR + '/db.sqlite3')
    cursor = con.cursor()
    lat_total = 0.0
    lng_total = 0.0
    data = cursor.execute(FETCH_LAT_LNG_SQL.format(tbl=tbl,
                                                   city=city,
                                                   timezone=timezone)).fetchall()

    for r in data:
        lat_total += _flt(r[1])
        lng_total += _flt(r[2])
    lat_avg = lat_total / len(data)
    lng_avg = lng_total / len(data)
    # sqlite3 doesn not have built-in functions for sqrt, power. I define them below
    con.create_function('power', 2, lambda base, power: base**power)
    con.create_function('sqrt', 1, lambda base: math.sqrt(base))
#    print (cursor.execute(CITIES_IN_DISTANCE_SQL.format(tbl=tbl,
#                                                 lat=lat_avg,
#                                                 lng=lng_avg,
#                                                 radius=radius)).fetchall())
    cities = []
    [cities.append(r[0]) for r in cursor.execute(CITIES_IN_DISTANCE_SQL.format(tbl=tbl,
                                                                           lat=lat_avg,
                                                                           lng=lng_avg,
                                                                           radius=radius)).fetchall()]
    return cities

# below is to load data to db, the process could be a few minutes, please be patient
import csv
INSERT_GEOLOCATION_SQL = """INSERT INTO {tbl} ({fltr_key})
                            VALUES ({fltr_value})
                            """
def import_from_csv_to_db():
    path = '/Users/Jerry/CA/'
    tbl = 'users_geolocation'
    cursor = connection.cursor()
    fields_list = []
    # for sqlite3
    [fields_list.append(r[1]) for r in cursor.execute(FIELDS_SQL.format(tbl=tbl)).fetchall()]
    with open(path+'CA.csv', 'r') as f:
        readCSV = csv.reader(f, delimiter=',')
        for row in readCSV:
            clear_row = []
            [clear_row.append(r.replace("'", "''")) for r in row]
            cursor.execute(INSERT_GEOLOCATION_SQL.format(tbl=tbl,
                                                         fltr_key=','.join(["""%s""" % (k) for k in fields_list]),
                                                         fltr_value=','.join(["""'%s'""" % (k) for k in clear_row])
                                                         ))

#=====================================================================================
# return a json with city name, lat, lng
def city_lat_lng():
    result = {}
    # change the path to your local
#    path = '/Users/Jerry/CA/'
    # file name is up to you
#    data = json.loads(open(path+'test.json', 'r').read())
    url = 'http://geogratis.gc.ca/services/geoname/en/geonames.json?lat=49.88805600000006&lon=-119.49555599999997&radius=50&theme=985'
    data = request_response(url)
    # parse data
    for r in data['items']:
        lat_lng = {}
        lat_lng.setdefault('latitude', r['latitude'])
        lat_lng.setdefault('longitude', r['longitude'])
        result.setdefault(r['name'], lat_lng)
    return result

from django.http import HttpResponse
import csv
import tempfile
from openpyxl import Workbook
from datetime import datetime
# this is to download csv
def jerry(request, client_attr, client_attr_2):
#    try:
#        int(client_attr)
#        int(client_attr_2)
#        resp = 'para_1: {para_1}<br>para_2: {para_2}'
#    except:
#        resp = 'Hello {para_1}<br>para_2: {para_2}'
#    return HttpResponse(resp.format(para_1=client_attr,
#                                    para_2=client_attr_2))
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Jerry_Django_%s.csv"'%datetime.strftime(datetime.now(), '%Y-%m-%d')
    writer = csv.writer(response)
    writer.writerow([client_attr, client_attr_2])
    return response

def jerry_2(request):
    current_namespace = request.resolver_match.namespace
    print(current_namespace)
    latest_location_list = GeoLocation.objects.order_by('id')[:5]
    # raise RuntimeError([r['name'] for r in latest_location_list.values()])
    context = {'latest_location_list': latest_location_list}
    return render(request, 'users/JERRY_2.html', context)

# this is to download xlsx
def jerry_3(request, client_attr, client_attr_2):
    file_name = 'Jerry_3_%s.xlsx' % datetime.strftime(datetime.now(), '%Y-%m-%d')
    tempfile_path = '%s/%s' % (tempfile.gettempdir(), file_name)
    # create a new XLSX template, it has one sheet by default
    xlsx = Workbook()
    #ws = xlsx.get_active_sheet()
    ws = xlsx[xlsx.sheetnames[0]]
    # title row
    row_id = 1
    ws['A%s' % row_id] = 'Personal Info'
    row_id += 1
    ws['A%s' % row_id] = 'Name'
    ws['B%s' % row_id] = 'Age'
    # starting row of data
    row_id = 3
    ws['A%s' % row_id] = '%s' % client_attr
    ws['B%s' % row_id] = '%s' % client_attr_2
    try:
        xlsx.save(tempfile_path)
        response = HttpResponse(open(tempfile_path, 'rb'), content_type='text/xlsx')
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    finally:
        os.unlink(tempfile_path)
    return response

import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
# this is to download pdf
def jerry_4(request):
    file_name = 'Jerry_4_%s.pdf' % datetime.strftime(datetime.now(), '%Y-%m-%d')
    # Create a file-like buffer to receive PDF data.
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)
    p.drawString(100, 10, "Hello Tommy.")
    # Close the PDF object cleanly
    try:
        p.showPage()
        p.save()
    finally:
        buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=file_name)
    
# this is to download pdf, the same as jerry_4()
def jerry_5(request):
    file_name = 'Jerry_5_%s.pdf' % datetime.strftime(datetime.now(), '%Y-%m-%d')
    tempfile_path = '%s/%s' % (tempfile.gettempdir(), file_name)
    p = canvas.Canvas(tempfile_path)
    p.drawString(10, 5, "Hello John.")
    # Close the PDF object cleanly
    try:
        p.showPage()
        p.save()
        response = HttpResponse(open(tempfile_path, 'rb'), content_type='text/pdf')
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    finally:
        os.unlink(tempfile_path)
    return response
